import datetime
import io
from decimal import Decimal

from openpyxl import Workbook


def _normalize_value(v):
    """Convert problematic types to Excel-friendly values.

    - timezone-aware datetimes: convert to UTC then remove tzinfo
    - Decimal: convert to float
    - None: return empty string
    """
    if v is None:
        return ""
    # datetime.datetime includes datetime.date as subclass; check datetime first
    if isinstance(v, datetime.datetime):
        if v.tzinfo is not None:
            try:
                v = v.astimezone(datetime.timezone.utc).replace(tzinfo=None)
            except Exception:
                # fallback: drop tzinfo
                v = v.replace(tzinfo=None)
        return v
    if isinstance(v, Decimal):
        try:
            return float(v)
        except Exception:
            return str(v)
    return v


def build_workbook_from_data(data: dict) -> io.BytesIO:
    """Construct an XLSX workbook from provided data dict and return BytesIO.

    Expected keys in data: contacts, organisations, roles, engagements,
    referencedata, sectors, contact_target_orgs, documents, engagement_documents
    Each value should be a list of dict-like rows (e.g., RealDictCursor results).
    """
    wb = Workbook()

    # Contacts sheet
    contacts = data.get("contacts") or []
    ws = wb.active
    ws.title = "Contacts"
    if contacts:
        headers = list(contacts[0].keys())
        ws.append(headers)
        for r in contacts:
            ws.append([_normalize_value(r.get(h)) for h in headers])
    else:
        ws.append(["contactid"])

    # Organisations
    organisations = data.get("organisations") or []
    ws = wb.create_sheet("Organisations")
    if organisations:
        headers = list(organisations[0].keys())
        ws.append(headers)
        for r in organisations:
            ws.append([_normalize_value(r.get(h)) for h in headers])
    else:
        ws.append(["orgid"])

    # Roles
    roles = data.get("roles") or []
    ws = wb.create_sheet("Roles")
    if roles:
        headers = list(roles[0].keys())
        ws.append(headers)
        for r in roles:
            ws.append([_normalize_value(r.get(h)) for h in headers])
    else:
        ws.append(["jobid"])

    # Engagements
    engagements = data.get("engagements") or []
    ws = wb.create_sheet("Engagements")
    if engagements:
        headers = list(engagements[0].keys())
        ws.append(headers)
        for r in engagements:
            ws.append([_normalize_value(r.get(h)) for h in headers])
    else:
        ws.append(["engagementlogid"])

    # ReferenceData
    referencedata = data.get("referencedata") or []
    ws = wb.create_sheet("ReferenceData")
    if referencedata:
        headers = list(referencedata[0].keys())
        ws.append(headers)
        for r in referencedata:
            ws.append([_normalize_value(r.get(h)) for h in headers])

    # Sectors
    sectors = data.get("sectors") or []
    ws = wb.create_sheet("Sectors")
    if sectors:
        headers = list(sectors[0].keys())
        ws.append(headers)
        for r in sectors:
            ws.append([_normalize_value(r.get(h)) for h in headers])

    # ContactTargetOrganisations
    contact_target_orgs = data.get("contact_target_orgs") or []
    ws = wb.create_sheet("ContactTargetOrganisations")
    if contact_target_orgs:
        headers = list(contact_target_orgs[0].keys())
        ws.append(headers)
        for r in contact_target_orgs:
            ws.append([_normalize_value(r.get(h)) for h in headers])

    # Documents
    documents = data.get("documents") or []
    ws = wb.create_sheet("Documents")
    if documents:
        headers = list(documents[0].keys())
        ws.append(headers)
        for r in documents:
            ws.append([_normalize_value(r.get(h)) for h in headers])

    # EngagementDocuments
    engagement_documents = data.get("engagement_documents") or []
    ws = wb.create_sheet("EngagementDocuments")
    if engagement_documents:
        headers = list(engagement_documents[0].keys())
        ws.append(headers)
        for r in engagement_documents:
            ws.append([_normalize_value(r.get(h)) for h in headers])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
